import openpyxl
"""
# 1.   WRITING SAME DATA IN ALL ROWS AND COLUMNS
file="C://Users//Reka//Desktop//New//selenium//writedatatoexcel.xlsx"
workbook=openpyxl.load_workbook(file)     # load /extract workbook from file
sheet=workbook["Data"]
# if u have only one sheet in ur workbook then u can write (sheet= workbook.active), to get active sheet from excel
# we have to decide how many rows and columns we are going to write 5 means 4 coulmns
for r in range (1,6):    # 6 means till 5 rows
    for c in range(1,4):      # 4 means 3 columns
        sheet.cell(r,c).value="Welcome"   # to set the value into cell
# create cell in excel sheet and assign value into excel sheet
# to save the data into file
workbook.save(file)
"""
# WRITE MULTIPLE DATA INTO EXCEL
file="C://Users//Reka//Desktop//New//selenium//writeDiffData.xlsx"
workbook=openpyxl.load_workbook(file)     # load /extract workbook from file
sheet=workbook["Sheet1"]
sheet.cell(1,1).value="Reka"   # to set the value into cell
sheet.cell(1,2).value=111
sheet.cell(1,3).value="India"
sheet.cell(2,1).value="Raja"
sheet.cell(2,2).value=222
sheet.cell(2,3).value="France"
sheet.cell(3,1).value="NV"
sheet.cell(3,2).value=333
sheet.cell(3,3).value="Pondy"
workbook.save(file)
