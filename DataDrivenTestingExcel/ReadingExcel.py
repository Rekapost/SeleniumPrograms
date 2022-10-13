import openpyxl
# file-->Workbook--->Sheets---->Rows---->Cell
file="C://Users//Reka//Desktop//New//selenium//DDTexcel.xlsx"    # File
workbook=openpyxl.load_workbook(file)     # load /extract workbook from file
sheet=workbook["Sheet1"]
# Loop statement to read all data( rows and columns) from Excel
# 2 loops to increment Rows and Increment columns
# first fine total no of rows and columns
# to find no of rows in excel
noofRows=sheet.max_row
# to find no of columns
noofColumns=sheet.max_column
#Reading rows and columns from excel   every row has multiple cells
for r in range(1,noofRows+1):
    for c in range(1,noofColumns+1):
        print(sheet.cell(r,c).value,end="    ")   # to print in tabular format
    print()  #  prints in set  set